import logging
import math
import os
from datetime import datetime

import pyodbc
from openpyxl import Workbook

# Connection String
connection_string = pyodbc.connect(
    'Driver={SQL Server};'
    'Server=.\sqlexpress;'
    'Database=DB_A4A307_Production_Migration_test;'
    'Trusted_Connection=yes;')

cursor = connection_string.cursor()

# Excel
wb = Workbook(write_only=True)
migration_ws = wb.create_sheet("Migration Result")
department_ws = wb.create_sheet("Department Status")
applicants_ws = wb.create_sheet("Applicants Status")


def backup_db( ):
    currentDirectory = os.getcwd()
    # print(currentDirectory)
    # print("Database backup started...")
    logging.info("Database backup started...")
    connection_string.autocommit = True
    backup_file = currentDirectory + "\\" + "db" + str(datetime.now().strftime("%d_%b_%I_%M_%p")) + ".bak"
    # sql = "BACKUP DATABASE [Admission2019] TO DISK = N" + backup_file
    query = "backup database [Admission2019] to disk = '%s'" % backup_file
    cursor.execute(query)
    cursor.commit()
    # connection_string.close()
    # cursor.execute(sql)
    # connection_string.autocommit = False
    logging.info("Find the backup file in " + backup_file)
    print("Database backup finished...")
    logging.info("Database backup finished...")


# Getting Applicants who did not cancelled admission and did not initiate Auto Migration OFF of Unit
def get_applicants(unit):
    logging.info(
        "Getting Applicants who did not cancelled admission and did not initiate Auto Migration OFF of Unit " + unit)
    try:
        query = "SELECT * FROM PassedApplicants" + unit + " WHERE [IsAdmissionCancelled] != 1 and [IsAutoMigrationOff] != 1 and UnitName = " + "'" + unit + "'"
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]
    except Exception as e:
        logging.info(e)


# get starting position of migration
def get_first_position(unit):
    logging.info("Getting 1st Applicant...")
    query = "SELECT MIN(Position) FROM PassedApplicants" + unit + " WHERE [IsAdmissionCancelled] != 1 and [IsAutoMigrationOff] != 1 and UnitName = " + "'" + unit + "'"
    cursor.execute(query)
    for row in cursor.fetchall():
        return row[0]


# get starting position of migration
def get_last_position(unit):
    logging.info("Getting 1st Applicant...")
    query = "SELECT MAX(Position) FROM PassedApplicants" + unit + " WHERE [IsAdmissionCancelled] != 1 and [IsAutoMigrationOff] != 1 and UnitName = " + "'" + unit + "'"
    cursor.execute(query)
    for row in cursor.fetchall():
        return row[0]


def get_allotted_subject_order(application_id, unit):
    logging.info("Getting order of allotted department...")
    query = "SELECT AllottedDepartmentOrder, AllottedDepartmentId FROM PassedApplicants" + unit + " WHERE Id =" + str(
        application_id)
    cursor.execute(query)
    for row in cursor.fetchall():
        return row[0], row[1]


def get_subject_choices_by_id(application_id):
    cursor.execute(
        "SELECT SubjectId, [Order], ApplicationId FROM Admission2019.SubjectChoices WHERE ApplicationId ='%s'" % application_id)
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def get_subject_id_by_order(application_table_id, order):
    cursor.execute("SELECT SubjectId FROM Admission2019.SubjectChoices WHERE ApplicationId = ? AND [Order] = ?",
                   application_table_id, order)
    for row in cursor.fetchall():
        return row[0]


def get_department_status_by_id(department_id, unit):
    query = "SELECT SeatStatus, TotalSeats, AllottedSeats, DepartmentName FROM Departments" + unit + " WHERE Id =" + str(
        department_id)
    cursor.execute(query)
    for row in cursor.fetchall():
        return row[0], row[1], row[2], row[3]


def get_departments(unit):
    query = "SELECT * FROM Departments " + unit + " WHERE SeatStatus = 1"
    cursor.execute(query)
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def allocate_subject(application_table_id, prev_department_order, prev_department_id, unit):
    choices = get_subject_choices_by_id(application_table_id)
    number_of_choices = len(choices)
    logging.info("No. of Choices: " + str(number_of_choices))
    x = datetime.now()
    order = 1
    while order <= number_of_choices and order < prev_department_order:
        subject_id = get_subject_id_by_order(application_table_id, order)
        seat_status, total_seats, allotted_seats, department_name = get_department_status_by_id(subject_id, unit)
        logging.info(
            str(order) + ": " + str(department_name) + " Total Seats: " + str(total_seats) + " Allotted Seats: " + str(
                allotted_seats))
        if seat_status is True and allotted_seats < total_seats and total_seats is not 0:
            if prev_department_id is not 0:
                logging.info("Previously allotted department id: " + str(prev_department_id))
                p_seat_status, p_total_seats, p_allotted_seats, p_department_name = get_department_status_by_id(
                    prev_department_id, unit)
                p_seat_status = True
                p_allotted_seats = p_allotted_seats - 1
                query = "UPDATE Departments" + unit + " SET [SeatStatus] =" + str(
                    p_seat_status) + " , [AllottedSeats] =" + str(p_allotted_seats) + " , [UpdatedDate] = " + "'" + str(
                    x.strftime("%d%b%I%M%p")) + "'" + " WHERE [Id] =" + str(prev_department_id)
                cursor.execute(query)
            query = "UPDATE PassedApplicants" + unit + " SET [AllottedDepartmentId] = " + str(
                subject_id) + " , [AllottedDepartment] =" + "'" + str(
                department_name) + "'" + " , [AllottedDepartmentOrder] = " + str(
                order) + " , [UpdatedDate] = " + "'" + str(
                x.strftime("%d%b%I%M%p")) + "'" + " WHERE [Id] = " + str(application_table_id)
            cursor.execute(query)
            allotted_seats = allotted_seats + 1
            if allotted_seats == total_seats:
                seat_status = 0
            else:
                seat_status = 1
            query = "UPDATE Departments" + unit + " SET [SeatStatus] =" + str(
                seat_status) + " , [AllottedSeats] =" + str(
                allotted_seats) + " , [UpdatedDate] =" + "'" + str(
                x.strftime("%d%b%I%M%p")) + "'" + " WHERE [Id] =" + str(
                subject_id)
            cursor.execute(query)
            return department_name
        else:
            order = order + 1
    return "No Department"


def get_applicant_id_by_position(applicant_position, unit):
    query = "SELECT Id FROM PassedApplicants" + unit + " WHERE UnitName = " + "'" + unit + "'" + " AND Position = " + str(
        applicant_position)
    cursor.execute(query)
    for row in cursor.fetchall():
        return row[0]


def write_department_data_to_excel(department_data):
    print("Writing department to excel......")
    logging.info("Writing department to excel......")
    # write header
    department_ws.append(
        ["Department", "Unit", "Total Seats", "Allotted Seats", "Seat Status", "UpdatedDate"])
    # write data
    for department in department_data:
        department_ws.append(
            [department[2], department[3], department[4], department[5], department[6], department[7]])


def write_migration_data_to_excel(applicants_data):
    logging.info("Writing migration data to excel......")
    print("Writing migration data to excel......")
    # write header
    migration_ws.append(
        ["Position", "Name", "ApplicationId", "Roll", "Unit", "Department", "Allotted Department Order", "Updated Date",
         "Auto Migration OFF"])
    data = list(applicants_data)
    # write data
    count = 0
    for applicant in data:
        migration_ws.append(
            [applicant[1], applicant[7], applicant[2], applicant[3], applicant[4], applicant[12],
             applicant[13], applicant[14], applicant[16]])
        count = count + 1


def execute_migration(unit_name):
    unit_name = unit_name.upper()

    drive = "F"
    os.chdir(drive + ':')  # change directory
    path = "Unit_" + unit_name + "_Migration" + str(datetime.now().strftime("_%d_%b_%I_%M_%p"))
    if not os.path.exists(path):
        os.makedirs(path)
    os.chdir(path)
    result = "Find migration results in " + str(drive.upper()) + ":\\" + str(path)
    logging.basicConfig(filename="info.log", format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S',
                        level=logging.INFO)

    if unit_name == 'A' or unit_name == 'B' or unit_name == 'C' or unit_name == 'D' or unit_name == 'E' or unit_name == 'F':
        backup_db()
        applicants = get_applicants(unit_name)
        no_of_applicants = len(applicants)
        logging.info("Total: " + str(no_of_applicants))

        position = get_first_position(unit_name)  # get starting position of migration
        logging.info("Starting Position: " + str(position))

        last_position = get_last_position(unit_name)  # get starting position of migration
        logging.info("Starting Position: " + str(position))

        while True:
            get_no_of_vacant_departments = "SELECT COUNT(SeatStatus) FROM Departments" + unit_name + " WHERE SeatStatus = 1 AND UnitName = " + "'" + unit_name + "'"
            cursor.execute(get_no_of_vacant_departments)
            for row in cursor.fetchall():
                no_of_vacant_departments = row[0]
            logging.info(no_of_vacant_departments)

            if position > last_position:
                logging.info("No applicant remains")
                break
            if no_of_vacant_departments is 0:
                logging.info("No vacant department remains")
                break

            logging.info(
                "------------------------------Migration Started for Position " + str(position) + "----------------------------------")
            applicant_id = get_applicant_id_by_position(position, unit_name)
            logging.info("Id: " + str(applicant_id))

            if applicant_id is not None:
                allotted_subject_order, allotted_subject_id = get_allotted_subject_order(applicant_id, unit_name)
                if allotted_subject_order is 0:
                    allotted_subject_order = math.inf
                    logging.info("No department is allotted yet")
                else:
                    logging.info("Order of currently allotted department: ")
                    logging.info(allotted_subject_order)
                    logging.info("currently allotted sub id " + str(allotted_subject_id))

                if applicant_id is not None and allotted_subject_order > 1:
                    department = allocate_subject(applicant_id, allotted_subject_order, allotted_subject_id, unit_name)
                    logging.info("Position " + str(position) + " " + department)
                    cursor.commit()  # Commit db changes

            position = position + 1
            no_of_applicants = no_of_applicants - 1

        logging.info("Exporting migration result into excel....")
        # get_applicants_query = " SELECT * FROM PassedApplicants WHERE IsAdmissionCancelled = 0 AND AllottedDepartment IS NOT NULL ORDER By Position asc"
        query = "SELECT * FROM PassedApplicants" + unit_name + " WHERE IsAdmissionCancelled = 0 AND AllottedDepartment IS NOT NULL and UnitName =" + "'" + unit_name + "'" + " ORDER By Position asc"
        migration_data = cursor.execute(query)

        # migration_data = cursor.execute(get_applicants_query)
        write_migration_data_to_excel(migration_data)

        logging.info("Exporting department status into excel....")
        # get_departments_query = " SELECT * FROM Departments"
        query = "SELECT * FROM Departments" + unit_name + " WHERE UnitName = " + "'" + unit_name + "'"
        department_data = cursor.execute(query)
        # department_data = cursor.execute(get_departments_query)
        write_department_data_to_excel(department_data)

        migration_result = "Migration.xlsx"
        wb.save(migration_result)
        logging.info("Find excel file into " + migration_result)
        result = "Migration is completed successfully! " + result
        parent_dir = os.path.dirname(os.getcwd())
        os.chdir(parent_dir)
        return result
    else:
        logging.warning("Invalid Unit Name: " + unit_name)
